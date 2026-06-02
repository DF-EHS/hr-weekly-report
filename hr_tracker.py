#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
hr_tracker.py - 人資部週追蹤報告產生器

讀取「資料/人資部週追蹤報告.xlsx」中的項目清單，
結合 Outlook 本週日誌，透過 Gemini 2.5 Flash 幕僚分析，
自動填入本週成果、達成率、RAG 狀態、卡關需求，
輸出 hr_weekly_report.html 供大老闆一目了然檢視。

使用方式：
  python hr_tracker.py               # 以今天推算本週期
  python hr_tracker.py 2026-05-21    # 指定週期內任一天
"""

import json
import re
import sys
import io
from datetime import date, datetime, timedelta
from pathlib import Path

# 排程執行時強制 stdout/stderr 使用 UTF-8
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

try:
    import requests
except ImportError:
    print("❌ 缺少套件：請先執行  pip install requests")
    sys.exit(1)

# ─────────────────────────────────────────────
# 套件載入檢查
# ─────────────────────────────────────────────
try:
    import openpyxl
except ImportError:
    print("❌ 缺少套件：請先執行  pip install openpyxl")
    sys.exit(1)

# 從 generate_report.py 重用共用函式（Outlook 讀取、Gemini 呼叫等）
SCRIPT_DIR = Path(__file__).parent
sys.path.insert(0, str(SCRIPT_DIR))
try:
    from generate_report import (
        load_config,
        get_week_range,
        get_emails,
        identify_sender,
        _call_gemini,
        fmt_date_zh,
    )
except ImportError as exc:
    print(f"❌ 無法載入 generate_report.py：{exc}")
    sys.exit(1)

# ─────────────────────────────────────────────
# 路徑常數
# ─────────────────────────────────────────────
EXCEL_PATH    = SCRIPT_DIR / "資料" / "人資部週追蹤報告.xlsx"
TEMPLATE_PATH = SCRIPT_DIR / "hr_tracker_template.html"
OUTPUT_PATH   = SCRIPT_DIR / "hr_weekly_report.html"

# Excel 負責人名稱 → Outlook 發信人識別名稱（對應 sender_mapping 輸出）
EXCEL_TO_OUTLOOK: dict[str, str] = {
    "婕誼": "Chiehyi",
    "詩紜": "An",
    "Bella": "Bella",
}

# RAG 四種狀態定義（用於 Gemini prompt 與 HTML 渲染）
RAG_META: dict[str, tuple[str, str]] = {
    "已完成": ("#1a8c6e", "✅ 已完成"),
    "正常推進": ("#28a745", "🟢 正常推進"),
    "待改善": ("#e6a800", "🟡 待改善"),
    "卡關": ("#dc3545", "🔴 卡關"),
    "本週無資料": ("#adb5bd", "⚪ 本週無資料"),
}

# 成員標籤顏色
OWNER_COLORS: dict[str, str] = {
    "婕誼":  "#6c757d",
    "詩紜":  "#2a7dd9",
    "Bella": "#7b5ea7",
    "柏羽":  "#c0392b",
    "信佳":  "#1a8c6e",
    "佑毓":  "#e07b39",
}


# ─────────────────────────────────────────────
# Excel 解析：讀取項目清單
# ─────────────────────────────────────────────
def read_excel_projects(excel_path: Path) -> list[dict]:
    """
    讀取 Excel 第一張工作表，從第 5 列（index 4）開始解析項目。
    合併儲存格的值以「向下延伸」方式處理。
    回傳項目清單，每筆為 dict。
    """
    wb = openpyxl.load_workbook(str(excel_path), data_only=True)
    ws = wb.active

    projects: list[dict] = []
    last_category = ""
    last_benefit  = ""

    for row_idx, row_cells in enumerate(ws.iter_rows(min_row=5), start=5):
        values = [c.value for c in row_cells]
        col_category, col_benefit, col_goal, col_owner, \
            col_result, col_rate, col_deadline, _col_rag, \
            col_blocker, col_benefit_actual, col_note = (
                values[i] if i < len(values) else None for i in range(11)
            )

        # 跳過完全空白列
        if all(v is None for v in (col_category, col_goal, col_owner)):
            continue

        # 合併儲存格：延伸上方非空值
        if col_category:
            last_category = str(col_category).strip()
        if col_benefit:
            last_benefit = str(col_benefit).strip()

        # 負責人為空 → 跳過（目標可為空，以項目名稱代替）
        if not col_owner:
            continue

        # 預計完成日格式化
        if isinstance(col_deadline, datetime):
            deadline_str = col_deadline.strftime("%Y/%m/%d")
        elif isinstance(col_deadline, str):
            deadline_str = col_deadline
        else:
            deadline_str = "—"

        # 基準達成率（Excel 已有的數字，0.0–1.0）
        base_rate = 0
        if isinstance(col_rate, (int, float)):
            base_rate = int(col_rate * 100) if col_rate <= 1.0 else int(col_rate)

        # 目標為空時用項目名稱代替
        goal_text = str(col_goal).strip() if col_goal else last_category

        projects.append({
            "category":       last_category,
            "benefit":        last_benefit,
            "goal":           goal_text,
            "owner_display":  str(col_owner).strip(),
            "outlook_name":   EXCEL_TO_OUTLOOK.get(str(col_owner).strip(), ""),
            "base_rate":      base_rate,
            "deadline":       deadline_str,
            "benefit_actual": str(col_benefit_actual).strip() if col_benefit_actual else "",
            "note":           str(col_note).strip() if col_note else "",
            "excel_row":      row_idx,   # 記錄 Excel 行號以便寫回
            # AI 填入欄位（初始為空）
            "weekly_result":  "",
            "ai_rate":        None,
            "rag":            "",
            "blocker":        "",
        })

    wb.close()
    return projects


# ─────────────────────────────────────────────
# Gemini：幕僚分析（每位成員一次呼叫）
# ─────────────────────────────────────────────
def call_gemini_tracker(
    api_key: str,
    member_display: str,
    weekly_logs: str,
    member_projects: list[dict],
) -> list[dict]:
    """
    請 Gemini 分析某成員的本週日誌，對應其負責的每個項目目標，
    回傳含 weekly_result / ai_rate / rag / blocker 的 JSON 陣列。
    """
    # 組建項目清單 JSON 給 Gemini 參考
    goals_for_prompt = [
        {"goal": p["goal"], "base_rate_pct": p["base_rate"]}
        for p in member_projects
    ]
    goals_json = json.dumps(goals_for_prompt, ensure_ascii=False, indent=2)

    prompt = (
        f"你是一位高階幕僚助理，負責協助分析人資部門同仁的週工作日誌。\n"
        f"同仁姓名：{member_display}\n\n"
        f"【本週工作日誌彙整】\n{weekly_logs[:8000]}\n\n"
        f"【請對應以下每個項目目標逐一分析】\n{goals_json}\n\n"
        "對每個項目，請依日誌內容填入以下欄位：\n"
        "1. weekly_result：50-100 字的本週具體進度成果（繁體中文）。"
        "若日誌中完全沒有提及此項目相關進度，請填「本週無實質進度」。\n"
        "2. ai_rate：整數 0-100，評估「此項目目前整體完成度」（參考 base_rate_pct 作為基準，"
        "本週若有推進可適度調高，若本週無進度維持基準值）。\n"
        "3. rag：必須從以下四選一，不可自行創造：「已完成」/ 「正常推進」/ 「待改善」/ 「卡關」\n"
        "4. blocker：若日誌提到報錯、伺服器失敗、權限問題、等待他人等阻礙，"
        "請精煉為 20 字以內描述；無卡關則填空字串。\n\n"
        "只回傳 JSON 陣列，不要任何說明文字或 markdown 標記。\n"
        '格式：[{"goal":"...","weekly_result":"...","ai_rate":80,"rag":"正常推進","blocker":""}]\n'
        "確保陣列長度與上方項目清單相同，順序對應。"
    )

    return _parse_tracker_response(api_key, prompt, member_projects)


def _parse_tracker_response(
    api_key: str, prompt: str, member_projects: list[dict]
) -> list[dict]:
    """直接呼叫 Gemini API 並解析陣列 JSON 回應"""
    import requests

    payload = {
        "contents": [{"parts": [{"text": prompt}]}],
        "generationConfig": {
            "temperature": 0.2,
            "maxOutputTokens": 2048,
            "thinkingConfig": {"thinkingBudget": 0},
        },
    }
    gemini_url = (
        "https://generativelanguage.googleapis.com/v1beta/models/"
        "gemini-2.5-flash:generateContent"
    )

    try:
        resp = requests.post(f"{gemini_url}?key={api_key}", json=payload, timeout=90)
        resp.raise_for_status()
        data = resp.json()

        parts = data["candidates"][0]["content"]["parts"]
        text  = "".join(p.get("text", "") for p in parts if not p.get("thought", False))
        text  = re.sub(r"^```(?:json)?\s*", "", text.strip())
        text  = re.sub(r"\s*```$", "", text).strip()

        # 嘗試解析 JSON 陣列
        m = re.search(r"\[[\s\S]*\]", text)
        if m:
            results = json.loads(m.group(0))
            if isinstance(results, list) and len(results) == len(member_projects):
                return results
            # 若長度不符，仍嘗試使用
            if isinstance(results, list):
                return results
    except Exception as exc:
        print(f"    ⚠️  Gemini 幕僚分析失敗：{exc}")

    # fallback：所有項目回傳空結果
    return [
        {
            "goal": p["goal"],
            "weekly_result": "（AI 分析失敗）",
            "ai_rate": p["base_rate"],
            "rag": "待改善",
            "blocker": "",
        }
        for p in member_projects
    ]


# ─────────────────────────────────────────────
# 合併 AI 結果至項目清單
# ─────────────────────────────────────────────
def merge_ai_results(projects: list[dict], ai_results: list[dict]) -> None:
    """將 Gemini 回傳的 ai_results 合併回 projects（依順序對應）"""
    for i, proj in enumerate(projects):
        if i >= len(ai_results):
            break
        res = ai_results[i]
        proj["weekly_result"] = str(res.get("weekly_result", "")).strip() or "本週無實質進度"
        proj["ai_rate"]       = int(res.get("ai_rate", proj["base_rate"]))
        proj["rag"]           = str(res.get("rag", "待改善")).strip()
        proj["blocker"]       = str(res.get("blocker", "")).strip()

        # 確保 rag 是合法值
        if proj["rag"] not in RAG_META:
            proj["rag"] = "待改善"


# ─────────────────────────────────────────────
# HTML 產生
# ─────────────────────────────────────────────
def _owner_badge(owner: str) -> str:
    return owner


def _rag_badge(rag: str) -> str:
    if not rag:
        rag = "本週無資料"
    color, label = RAG_META.get(rag, ("#adb5bd", f"⚪ {rag}"))
    return f'<span class="rag-badge" style="background:{color}">{label}</span>'


def _rate_bar(rate: int | None, base_rate: int) -> str:
    """產生可編輯的進度條 HTML，顯示 AI 評估達成率與原始基準"""
    if rate is None:
        rate = base_rate
    color = "#1a8c6e" if rate >= 100 else "#28a745" if rate >= 80 else \
            "#e6a800" if rate >= 60 else "#dc3545"
    return (
        f'<div class="rate-wrap">'
        f'<div class="rate-bar" style="width:{min(rate,100)}%;background:{color}"></div>'
        f'<input class="rate-input" type="number" min="0" max="100" value="{rate}" '
        f'oninput="updateRate(this)">'
        f'</div>'
    )


def build_table_rows(projects: list[dict]) -> str:
    if not projects:
        return '<tr><td colspan="9" class="no-data">尚無項目資料</td></tr>'

    rows_html: list[str] = []

    # 依照 category 分組，計算 rowspan
    from itertools import groupby
    groups = []
    for cat, items in groupby(projects, key=lambda p: p["category"]):
        items_list = list(items)
        groups.append((cat, items_list))

    for cat, items_list in groups:
        for idx, p in enumerate(items_list):
            is_first = (idx == 0)
            rowspan  = len(items_list)

            # 整列 CSS class
            rag_class = {
                "已完成": "row-done",
                "正常推進": "row-green",
                "待改善": "row-yellow",
                "卡關": "row-red",
            }.get(p["rag"], "row-none")

            row = f'<tr class="{rag_class}">\n'

            # 項目名稱（僅第一行用 rowspan）
            if is_first:
                row += (
                    f'  <td class="cat-cell" rowspan="{rowspan}">'
                    f'<span class="cat-label">{cat}</span></td>\n'
                )

            # 其餘欄位
            goal_html    = p["goal"].replace("\n", "<br>")
            benefit_html = p["benefit"].replace("\n", "<br>") if p["benefit"] else "—"
            note_html    = p["note"].replace("\n", "<br>") if p["note"] else "—"
            blocker_html = (
                f'<span class="blocker-text">{p["blocker"]}</span>'
                if p["blocker"] else '<span class="no-blocker">—</span>'
            )

            row += (
                f'  <td>{benefit_html}</td>\n'
                f'  <td class="goal-cell">{goal_html}</td>\n'
                f'  <td class="owner-cell">{_owner_badge(p["owner_display"])}</td>\n'
                f'  <td class="result-cell">{p["weekly_result"] or "—"}</td>\n'
                f'  <td class="rate-cell">{_rate_bar(p["ai_rate"], p["base_rate"])}</td>\n'
                f'  <td class="deadline-cell">{p["deadline"]}</td>\n'
                f'  <td class="rag-cell">{_rag_badge(p["rag"])}</td>\n'
                f'  <td class="blocker-cell">{blocker_html}</td>\n'
                f'</tr>\n'
            )
            rows_html.append(row)

    return "\n".join(rows_html)


def build_summary_stats(projects: list[dict]) -> dict:
    """計算摘要統計"""
    tracked = [p for p in projects if p["rag"]]
    done_n    = sum(1 for p in tracked if p["rag"] == "已完成")
    green_n   = sum(1 for p in tracked if p["rag"] == "正常推進")
    yellow_n  = sum(1 for p in tracked if p["rag"] == "待改善")
    red_n     = sum(1 for p in tracked if p["rag"] == "卡關")
    none_n    = len(projects) - len(tracked)
    total_n   = len(projects)

    return {
        "total":  total_n,
        "done":   done_n,
        "green":  green_n,
        "yellow": yellow_n,
        "red":    red_n,
        "none":   none_n,
    }


# ─────────────────────────────────────────────
# 匯出 Excel（AI 結果寫回範本，另存新檔）
# ─────────────────────────────────────────────
def export_excel(projects: list[dict], ref_date=None, week_start=None, week_end=None) -> tuple[Path, str]:
    """
    將 AI 分析結果寫入 Excel 副本並回傳 (檔案路徑, base64字串)。
    原始範本不會被覆蓋，另存為「人資部週追蹤報告(MMDD~MMDD).xlsx」。
    若外部傳入 week_start/week_end，則直接使用；否則依 ref_date 自行推算。
    """
    import shutil
    import base64
    from datetime import date as _date, timedelta
    from openpyxl.cell.cell import MergedCell as _MergedCell
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    FONT_NAME = "Microsoft JhengHei"
    _GRAY   = Side(style="thin", color="CCCCCC")
    _BORDER = Border(left=_GRAY, right=_GRAY, top=_GRAY, bottom=_GRAY)

    # RAG → (背景色, 文字色)
    _RAG_COLOR = {
        "已完成":  ("E6F4EA", "137333"),
        "正常推進": ("E8F0FE", "1A73E8"),
        "待改善":  ("FEF7E0", "B06000"),
        "卡關":    ("FCE8E6", "C5221F"),
    }

    def _safe_write(ws, row, col, value):
        """安全寫入：若為合併儲存格的從屬格，則找到主儲存格寫入"""
        cell = ws.cell(row=row, column=col)
        if isinstance(cell, _MergedCell):
            for rng in ws.merged_cells.ranges:
                if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
                    cell = ws.cell(rng.min_row, rng.min_col)
                    break
        cell.value = value

    def _style_cell(cell, size=10, bold=False, halign="left", valign="center",
                    wrap=True, border=True, fill_hex=None, color_hex=None):
        """對單一 Cell 套用字體／對齊／框線／填色"""
        cell.font = Font(
            name=FONT_NAME, size=size, bold=bold,
            color=color_hex if color_hex else "000000"
        )
        cell.alignment = Alignment(
            horizontal=halign, vertical=valign,
            wrap_text=wrap
        )
        if border:
            cell.border = _BORDER
        if fill_hex:
            cell.fill = PatternFill("solid", fgColor=fill_hex)

    if ref_date is None:
        ref_date = _date.today()

    # 實際週期由外部傳入，避免重複推算導致屏期錯位
    if week_start is None or week_end is None:
        days_since_wed = (ref_date.weekday() - 2) % 7
        week_start = ref_date - timedelta(days=days_since_wed)
        week_end   = week_start + timedelta(days=6)

    date_range   = f"{week_start.strftime('%m%d')}~{week_end.strftime('%m%d')}"
    out_filename = f"人資部週追蹤報告({date_range}).xlsx"

    # 儲存到本機「資料」資料夾，方便集中管理後手動或自動上傳
    out_path = EXCEL_PATH.parent / out_filename

    # 複製範本，保留格式
    shutil.copy(str(EXCEL_PATH), str(out_path))

    wb = openpyxl.load_workbook(str(out_path))
    ws = wb.active

    # ── 1. 寫入 AI 分析資料 ──────────────────────
    # E 欄（本週成果）在範本中是合併儲存格（如 E5:E8），
    # 同一合併範圍內的多個子項目必須拼接後一起寫入主格。
    # 先建立「E 主格 row → 屬於該合併範圍的 projects」對照表
    def _e_main_row(ws, row):
        """回傳 row 所在 E 欄合併範圍的起始 row（若無合併則回傳 row 本身）"""
        for rng in ws.merged_cells.ranges:
            if rng.min_col <= 5 <= rng.max_col and rng.min_row <= row <= rng.max_row:
                return rng.min_row
        return row

    # 依「E 主格 row」分組
    from collections import defaultdict
    e_groups: dict[int, list[dict]] = defaultdict(list)
    for p in projects:
        row_idx = p.get("excel_row")
        if not row_idx:
            continue
        e_groups[_e_main_row(ws, row_idx)].append(p)

    for main_row, group in e_groups.items():
        # 本週成果：將同一合併格內多個子項目的成果逐條拼接
        parts = []
        for p in group:
            result = (p.get("weekly_result") or "").strip()
            if result and result != "本週無日誌提交":
                goal_short = p.get("goal", "")[:20].replace("\n", " ")
                parts.append(f"【{goal_short}】{result}")
        combined = "\n".join(parts) if parts else (group[0].get("weekly_result", "") if group else "")
        ws.cell(main_row, 5).value = combined

    # F/H/I 欄各列獨立寫入（無合併，可直接 _safe_write）
    for p in projects:
        row_idx = p.get("excel_row")
        if not row_idx:
            continue
        ai_rate = p.get("ai_rate")
        if ai_rate is not None:
            _safe_write(ws, row_idx, 6, ai_rate / 100)   # 達成率% 存 0~1
        _safe_write(ws, row_idx, 8, p.get("rag", ""))
        _safe_write(ws, row_idx, 9, p.get("blocker", "") or "")

    # ── 2. 欄寬設定 ──────────────────────────────
    COL_WIDTHS = {
        1: 14,   # A 項目名稱
        2: 18,   # B 預期效益
        3: 30,   # C 目標
        4: 10,   # D 負責人
        5: 44,   # E 本週成果（最寬）
        6: 10,   # F 達成率%
        7: 13,   # G 預計完成日
        8: 12,   # H RAG
        9: 22,   # I 卡關/需求
        10: 16,  # J 實際效益
        11: 35,  # K 說明
    }
    for col_num, width in COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col_num)].width = width

    # ── 3. 標題列（第 4 列）樣式 ─────────────────
    ws.row_dimensions[4].height = 28
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=4, column=col)
        if isinstance(cell, _MergedCell):
            continue
        _style_cell(cell, size=11, bold=True, halign="center")

    # ── 4. 資料列（第 5 列起）樣式 ───────────────
    # 置中欄：D=負責人(4)  F=達成率%(6)  G=預計完成日(7)  H=RAG(8)
    CENTER_COLS = {4, 6, 7, 8}

    for row_idx in range(5, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 72
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col)
            if isinstance(cell, _MergedCell):
                continue
            halign = "center" if col in CENTER_COLS else "left"

            # RAG 欄特別處理顏色
            if col == 8 and cell.value in _RAG_COLOR:
                bg, fg = _RAG_COLOR[cell.value]
                _style_cell(cell, halign="center", fill_hex=bg, color_hex=fg, bold=True)
            else:
                _style_cell(cell, halign=halign)

            # 達成率%：百分比格式
            if col == 6 and isinstance(cell.value, (int, float)):
                cell.number_format = "0%"

    wb.save(str(out_path))
    wb.close()

    with open(str(out_path), "rb") as f:
        b64 = base64.b64encode(f.read()).decode("utf-8")

    return out_path, b64


# ─────────────────────────────────────────────
# 主程式
# ─────────────────────────────────────────────
def main() -> None:
    print("=" * 54)
    print("  人資部週追蹤報告產生器")
    print("=" * 54)
    print()

    # ── 1. 解析命令列參數（可指定日期） ──────────
    ref_date: date | None = None
    if len(sys.argv) >= 2:
        try:
            ref_date = datetime.strptime(sys.argv[1], "%Y-%m-%d").date()
            print(f"📌 指定日期：{sys.argv[1]}")
        except ValueError:
            print(f"⚠️  日期格式錯誤（應為 YYYY-MM-DD），改用今天")

    # ── 2. 載入設定 ────────────────────────────
    config  = load_config()
    api_key = config.get("gemini_api_key", "")
    if not api_key or api_key == "YOUR_GEMINI_API_KEY":
        print("❌ 請先在 config.json 中填入有效的 gemini_api_key")
        sys.exit(1)

    # ── 3. 計算週期 ────────────────────────────
    week_start, week_end = get_week_range(ref_date)
    week_label = f"{fmt_date_zh(week_start)} ~ {fmt_date_zh(week_end)}"
    print(f"📅 本週週期：{week_label}")
    print()

    # ── 4. 讀取 Excel 項目清單 ─────────────────
    print(f"📊 讀取 Excel：{EXCEL_PATH.name}")
    if not EXCEL_PATH.exists():
        print(f"❌ 找不到 Excel 檔案：{EXCEL_PATH}")
        sys.exit(1)
    all_projects = read_excel_projects(EXCEL_PATH)
    print(f"   共讀取 {len(all_projects)} 個項目")
    print()

    # ── 5. 讀取 Outlook 本週信件 ───────────────
    folder_path = config.get("outlook_folder_path", "")
    days_back   = int(config.get("days_to_look_back", 30))
    print(f"📬 讀取 Outlook（近 {days_back} 天）...")
    all_emails = get_emails(config["email_subjects"], days_back, folder_path)
    print(f"   共找到 {len(all_emails)} 封信件")

    # 識別發信人
    for e in all_emails:
        e["member"] = identify_sender(
            e["sender_name"], e["sender_email"], config["sender_mapping"]
        )

    # 只取本週信件（週三 ~ 下週二）
    week_emails = [
        e for e in all_emails
        if week_start <= e["date_obj"] <= week_end
    ]
    print(f"   本週（{week_start} ~ {week_end}）信件：{len(week_emails)} 封")
    print()

    # ── 6. 依成員分組週日誌 ────────────────────
    member_logs: dict[str, list[str]] = {}
    for e in week_emails:
        m = e["member"]
        # 只保留被 EXCEL_TO_OUTLOOK 對應到的成員
        if m not in EXCEL_TO_OUTLOOK.values():
            continue
        if m not in member_logs:
            member_logs[m] = []
        header = f"=== {e['date']} | {e['subject']} ===\n"
        member_logs[m].append(header + e["body"])

    # ── 7. Gemini 幕僚分析（每位追蹤成員一次） ──
    print("━" * 54)
    print("  [幕僚分析] 對應項目目標...")
    print("━" * 54)

    # 建立「outlook_name → projects」對照
    for outlook_name in EXCEL_TO_OUTLOOK.values():
        member_projects = [
            p for p in all_projects if p["outlook_name"] == outlook_name
        ]
        if not member_projects:
            continue

        # 取得顯示名稱（Excel 中的名字）
        display_name = next(
            k for k, v in EXCEL_TO_OUTLOOK.items() if v == outlook_name
        )

        logs_combined = "\n\n".join(member_logs.get(outlook_name, []))
        if not logs_combined.strip():
            print(f"  ⏭️  [{display_name}] 本週無日誌，跳過 AI 分析（維持 Excel 基準值）")
            # 填入「本週無資料」
            for p in member_projects:
                p["weekly_result"] = "本週無日誌提交"
                p["ai_rate"]       = p["base_rate"]
                p["rag"]           = "本週無資料"
                p["blocker"]       = ""
            continue

        print(f"  🤖 [{display_name}] 共 {len(member_logs.get(outlook_name, []))} 封日誌"
              f"，分析 {len(member_projects)} 個項目目標...")
        ai_results = call_gemini_tracker(
            api_key, display_name, logs_combined, member_projects
        )
        merge_ai_results(member_projects, ai_results)

        for p in member_projects:
            rag_icon = {"已完成": "✅", "正常推進": "🟢", "待改善": "🟡", "卡關": "🔴"}.get(
                p["rag"], "⚪"
            )
            print(f"    {rag_icon}  {p['goal'][:30]}  → {p['ai_rate']}%  {p['rag']}")

    print()

    # ── 8. 計算摘要統計 ────────────────────────
    stats = build_summary_stats(all_projects)

    # ── 9. 讀取 HTML 範本並填充 ─────────────────
    if not TEMPLATE_PATH.exists():
        print(f"❌ 找不到 HTML 範本：{TEMPLATE_PATH}")
        sys.exit(1)

    with open(TEMPLATE_PATH, "r", encoding="utf-8") as f:
        html = f.read()

    replacements = {
        "{{WEEK_LABEL}}":      week_label,
        "{{GENERATED_AT}}":    datetime.now().strftime("%Y-%m-%d %H:%M"),
        "{{TOTAL_COUNT}}":     str(stats["total"]),
        "{{DONE_COUNT}}":      str(stats["done"]),
        "{{GREEN_COUNT}}":     str(stats["green"]),
        "{{YELLOW_COUNT}}":    str(stats["yellow"]),
        "{{RED_COUNT}}":       str(stats["red"]),
        "{{NONE_COUNT}}":      str(stats["none"]),
        "{{TABLE_ROWS}}":      build_table_rows(all_projects),
    }
    for k, v in replacements.items():
        html = html.replace(k, v)

    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        f.write(html)

    print(f"✅ 完成！已產生：{OUTPUT_PATH}")
    print(f"   項目總計：{stats['total']} 個")
    print(f"   ✅ 已完成 {stats['done']}  🟢 正常推進 {stats['green']}"
          f"  🟡 待改善 {stats['yellow']}  🔴 卡關 {stats['red']}")


if __name__ == "__main__":
    main()
